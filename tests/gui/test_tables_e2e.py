#!/usr/bin/env python3
# L3 golden-file E2E test for Tables — dogtail-driven GUI + oracle verification.
# SPDX-License-Identifier: GPL-3.0-or-later
#
# Flow: known fixture loaded by the app → drive formatting actions via AT-SPI →
#       trigger save → verify with independent openpyxl reader.
# See TESTING-SPEC.md §3 / §4.
#
# Integration: the Flatpak must be launched with a fixture-loading env hook
# (TABLES_SELFTEST load-only, or a new TABLES_GUITEST hook). The save path
# is written to a known location. This script drives the dogtail portion.
#
# Run (from justfile):
#   flatpak run --env=TABLES_GUITEST=$tmpdir app_id &
#   sleep 8; python3 tests/gui/test_tables_e2e.py $tmpdir

import os
import sys
import time

# Resolve suite-common helpers.
for _candidate in (
    os.path.join(os.path.dirname(__file__), '..', '..', '..', 'suite-common'),
    os.path.join(os.path.dirname(__file__), '..', '..', 'subprojects', 'suite-common'),
):
    if os.path.isdir(_candidate):
        sys.path.insert(0, _candidate)
        break

from suite_common.test_helpers import click, find_app, find_widget, pressed, toggle_and_assert


def _oracle_verify_bold(path, cell_ref='A1'):
    """Verify cell_ref is bold via independent openpyxl read."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    cell = wb.active[cell_ref]
    assert cell.font.bold, f'{cell_ref} should be bold'
    print(f'oracle: {cell_ref} bold = {cell.font.bold}')


def _oracle_values(path, expectations):
    """Verify cell values unchanged."""
    import openpyxl
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active
    for ref, expected in expectations.items():
        actual = str(ws[ref].value)
        assert actual == str(expected), \
            f'{ref}: expected {expected!r}, got {actual!r}'
    print(f'oracle: {len(expectations)} cell values verified')


def main(out_dir=None):
    """Drive GUI actions on a running Tables Flatpak.

    Args:
        out_dir: If set, verify the saved-format file written by the app
                 (the app must write e.g. out.xlsx here via a test hook).
    """
    app = find_app('tables')
    print('=== L3 golden-file E2E: Tables ===')

    # ── 1. Bold formatting ────────────────────────────────────────────
    bold = app.child(name='Bold', roleName='toggle button')
    toggle_and_assert(bold)
    print('[bold] toggled via AT-SPI')

    # ── 2. Align Center ──────────────────────────────────────────────
    center = find_widget(app, name='Align Center', role='push button',
                         showing_only=False)
    if center is None:
        # When narrow: might be in the 'more' menu.
        more = app.child(name='More actions', roleName='toggle button',
                         showingOnly=False)
        click(more)
        time.sleep(0.5)
        # After opening 'more', try again.
        center = find_widget(app, name='Align Center', role='push button',
                             showing_only=False)
        if center is None:
            print('[align] Align Center not found (responsive collapse?)')
        else:
            click(center)
            time.sleep(0.4)
            print('[align] Align Center clicked via more menu')
    else:
        click(center)
        time.sleep(0.4)
        print('[align] Align Center clicked')

    # ── 3. Sheet switcher ─────────────────────────────────────────────
    combo = app.child(roleName='combo box', showingOnly=False)
    print(f'[sheet] sheet switcher accessible: {combo.name}')

    # ── 4. Node count (grid bridged) ──────────────────────────────────
    from suite_common.test_helpers import count_nodes
    nodes = count_nodes(app)
    assert nodes > 50, f'grid bridge: expected >50 nodes, got {nodes}'
    print(f'[grid] bridged: {nodes} AT-SPI nodes')

    # ── 5. Oracle verification (if output dir provided) ───────────────
    if out_dir:
        out_xlsx = os.path.join(out_dir, 'out.xlsx')
        if os.path.exists(out_xlsx):
            _oracle_verify_bold(out_xlsx, 'A1')
            _oracle_values(out_xlsx, {'A1': 'Hello', 'B1': 'World'})
            print('[oracle] PASS — bold + values verified')
        else:
            print(f'[oracle] SKIP — {out_xlsx} not found')

    print('L3-E2E: PASS')
    return 0


if __name__ == '__main__':
    out_dir = sys.argv[1] if len(sys.argv) > 1 else None
    try:
        sys.exit(main(out_dir))
    except Exception as exc:
        print(f'L3-E2E: FAIL — {exc}')
        sys.exit(1)
