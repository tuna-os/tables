# fileio.py — spreadsheet import/export adapters (in-process Python libraries).
# SPDX-License-Identifier: GPL-3.0-or-later
#
# Sheets are (name, rows_2d, styles) triples, where styles maps a cell ref
# ("A1") to a Jspreadsheet/CSS style string ("font-weight:bold;text-align:center").
# xlsx carries styles (openpyxl); csv/ods carry values only.

import csv
import os


def _ext(path):
    return os.path.splitext(path)[1].lower()


# ----- read -----------------------------------------------------------------

def read_spreadsheet(path):
    """Return a list of (sheet_name, rows_2d, styles) triples."""
    ext = _ext(path)
    if ext == '.csv':
        with open(path, newline='', encoding='utf-8') as fh:
            return [('Sheet 1', [list(r) for r in csv.reader(fh)], {})]
    if ext == '.xlsx':
        return _read_xlsx(path)
    if ext == '.ods':
        return _read_ods(path)
    raise ValueError(f'Unsupported format: {ext}')


def _cell_css(cell):
    parts = []
    font = cell.font
    if font and font.bold:
        parts.append('font-weight:bold')
    if font and font.italic:
        parts.append('font-style:italic')
    if font and font.color and isinstance(getattr(font.color, 'rgb', None), str):
        rgb = font.color.rgb
        if rgb[-6:] not in ('000000',):
            parts.append('color:#' + rgb[-6:])
    fill = cell.fill
    if fill and getattr(fill, 'fgColor', None) and isinstance(getattr(fill.fgColor, 'rgb', None), str):
        rgb = fill.fgColor.rgb
        if fill.patternType == 'solid' and rgb[-6:] not in ('000000',):
            parts.append('background-color:#' + rgb[-6:])
    if cell.alignment and cell.alignment.horizontal:
        parts.append('text-align:' + cell.alignment.horizontal)
    return ';'.join(parts)


def _read_xlsx(path):
    from openpyxl import load_workbook
    wb = load_workbook(path, data_only=True)
    sheets = []
    for ws in wb.worksheets:
        rows, styles = [], {}
        for row in ws.iter_rows():
            out = []
            for cell in row:
                out.append('' if cell.value is None else cell.value)
                css = _cell_css(cell)
                if css:
                    styles[cell.coordinate] = css
            rows.append(out)
        sheets.append((ws.title, rows, styles))
    return sheets or [('Sheet 1', [[]], {})]


def _read_ods(path):
    from odf.opendocument import load
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    doc = load(path)
    sheets = []
    for table in doc.spreadsheet.getElementsByType(Table):
        name = table.getAttribute('name') or 'Sheet'
        rows = []
        for tr in table.getElementsByType(TableRow):
            row = []
            for tc in tr.getElementsByType(TableCell):
                repeat = int(tc.getAttribute('numbercolumnsrepeated') or 1)
                text = ''.join(str(p) for p in tc.getElementsByType(P))
                row.extend([text] * repeat)
            rows.append(row)
        sheets.append((name, rows, {}))
    return sheets or [('Sheet 1', [[]], {})]


# ----- write ----------------------------------------------------------------

def write_spreadsheet(path, sheets):
    """sheets: list of (name, rows) or (name, rows, styles)."""
    ext = _ext(path)
    sheets = [_norm(s) for s in sheets]
    if ext == '.csv':
        _, rows, _ = sheets[0]
        with open(path, 'w', newline='', encoding='utf-8') as fh:
            csv.writer(fh, lineterminator='\n').writerows(rows)
        return
    if ext == '.xlsx':
        return _write_xlsx(path, sheets)
    if ext == '.ods':
        return _write_ods(path, sheets)
    raise ValueError(f'Unsupported format: {ext}')


def _norm(sheet):
    name, rows = sheet[0], sheet[1]
    styles = sheet[2] if len(sheet) > 2 else {}
    return name, rows, (styles or {})


def _coerce(value):
    if value is None or value == '':
        return None
    if isinstance(value, (int, float)):
        return value
    s = str(value)
    try:
        return int(s)
    except ValueError:
        pass
    try:
        return float(s)
    except ValueError:
        return s


def _parse_css(css):
    out = {}
    for decl in (css or '').split(';'):
        if ':' in decl:
            k, v = decl.split(':', 1)
            out[k.strip().lower()] = v.strip().lower()
    return out


def _hex(value):
    if not value:
        return None
    return value.lstrip('#').upper()[:6] or None


def _write_xlsx(path, sheets):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    wb = Workbook()
    wb.remove(wb.active)
    for name, rows, styles in sheets:
        ws = wb.create_sheet(title=(name or 'Sheet')[:31])
        for row in rows:
            ws.append([_coerce(c) for c in row])
        for ref, css in styles.items():
            try:
                cell = ws[ref]
            except Exception:  # noqa: BLE001
                continue
            d = _parse_css(css)
            font_kw = {}
            if 'bold' in d.get('font-weight', ''):
                font_kw['bold'] = True
            if 'italic' in d.get('font-style', ''):
                font_kw['italic'] = True
            if d.get('color'):
                font_kw['color'] = _hex(d['color'])
            if font_kw:
                cell.font = Font(**font_kw)
            bg = _hex(d.get('background-color'))
            if bg:
                cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type='solid')
            if d.get('text-align'):
                cell.alignment = Alignment(horizontal=d['text-align'])
    wb.save(path)


def _write_ods(path, sheets):
    from odf.opendocument import OpenDocumentSpreadsheet
    from odf.table import Table, TableRow, TableCell
    from odf.text import P
    doc = OpenDocumentSpreadsheet()
    for name, rows, _styles in sheets:
        table = Table(name=name or 'Sheet')
        for row in rows:
            tr = TableRow()
            for cell in row:
                tc = TableCell()
                tc.addElement(P(text='' if cell is None else str(cell)))
                tr.addElement(tc)
            table.addElement(tr)
        doc.spreadsheet.addElement(table)
    doc.save(path)
